import pandas as pd
from openpyxl import load_workbook
from openai import OpenAI
import json
import os
from dotenv import load_dotenv

load_dotenv()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


def get_column_mapping_from_gpt(columns: list) -> dict:
    """ให้ GPT สร้าง mapping จากชื่อ column ไทยเป็นอังกฤษ"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": """คุณเป็น expert ในการแปลงชื่อ column จากภาษาไทยเป็นภาษาอังกฤษ

กฎสำคัญ (ห้ามละเมิดเด็ดขาด):
1. column ที่หมายถึง "เดือน" หรือ "วันที่" หรือ "ช่วงเวลา" → ต้องชื่อ "date" เท่านั้น ห้ามใส่ปีหรือข้อมูลอื่นต่อท้าย
2. column ที่หมายถึง "ประเภทรายการ" หรือ "รายละเอียด" หรือ column ที่เก็บค่า budget/spent/remaining → ต้องชื่อ "details" เท่านั้น
3. ใช้ snake_case เท่านั้น (เช่น general_expense, asset_firewall)
4. ชื่อต้องสั้น กระชับ เข้าใจง่าย ไม่เกิน 60 ตัวอักษร
5. ถ้าเป็นภาษาอังกฤษอยู่แล้ว ให้แปลงเป็น snake_case
6. ตอบเป็น JSON object เท่านั้น ไม่ต้องอธิบาย
7. format: {"ชื่อเดิม": "ชื่อใหม่", ...}"""
            },
            {
                "role": "user",
                "content": f"แปลงชื่อ columns เหล่านี้เป็นภาษาอังกฤษ:\n{json.dumps(columns, ensure_ascii=False)}"
            }
        ]
    )

    result = response.choices[0].message.content.strip()
    if result.startswith("```"):
        lines = result.split("\n")
        lines = [line for line in lines if not line.startswith("```")]
        result = "\n".join(lines).strip()

    mapping = json.loads(result)

    # safety net: ถ้า GPT ยังแปลงผิด ให้ force override
    for orig, renamed in list(mapping.items()):
        orig_lower = orig.lower()
        # column ที่มีคำว่าเดือน/วันที่ → date
        if any(kw in orig_lower for kw in ["เดือน", "วันที่", "month", "date"]):
            if renamed != "date":
                print(f"   ⚠️ force override: '{renamed}' → 'date'")
                mapping[orig] = "date"
        # column ที่มีคำว่าประเภท/รายละเอียด → details
        if any(kw in orig_lower for kw in ["ประเภท", "รายละเอียด", "type", "detail"]):
            if renamed != "details":
                print(f"   ⚠️ force override: '{renamed}' → 'details'")
                mapping[orig] = "details"

    return mapping


def get_data_mapping_from_gpt(values: list) -> dict:
    """ให้ GPT สร้าง mapping สำหรับข้อมูลในแต่ละ cell"""

    # กรอง: เอาเฉพาะ string สั้น (ชื่อเดือน/ประเภท) ไม่เกิน 50 chars
    # ค่าที่ยาวกว่านี้มักเป็น cell หมายเหตุ ไม่ต้องแปลง
    values = [v for v in values if isinstance(v, str) and len(v) <= 50]
    if not values:
        return {}

    # แบ่ง batch ละ 50 values เพื่อไม่เกิน token limit
    results = {}
    batch_size = 50
    for i in range(0, len(values), batch_size):
        batch = values[i:i + batch_size]
        try:
            batch_result = _get_data_mapping_batch(batch)
            results.update(batch_result)
        except Exception as e:
            print(f"   ⚠️ batch {i//batch_size + 1} failed: {e}")
    return results


def _get_data_mapping_batch(values: list) -> dict:
    """แปลง batch เดียว"""
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        max_tokens=2000,
        messages=[
            {
                "role": "system",
                "content": """แปลงข้อมูลภาษาไทยเป็นภาษาอังกฤษ

กฎ:
1. เดือนไทย -> YYYY-MM format
   - ปีไทย (พ.ศ.) แปลงเป็น ค.ศ. โดย ค.ศ. = พ.ศ. - 543
   - เลขปีย่อ 2 หลัก เช่น 65, 66, 67 = พ.ศ. 2565, 2566, 2567
   - ดังนั้น 65 = 2022, 66 = 2023, 67 = 2024, 68 = 2025

   ชื่อเดือน:
   - มค=01, กพ=02, มีค=03, เมย=04, พค=05, มิย=06
   - กค=07, สค=08, กย=09, ตค=10, พย=11, ธค=12

   ตัวอย่าง: ตค65 -> 2022-10, มค66 -> 2023-01, กย67 -> 2024-09

2. เดือนอังกฤษ -> YYYY-MM format
   - Oct 2022 -> 2022-10, Jan 2023 -> 2023-01

3. ประเภทรายการ -> English
   - ยอดงบประมาณ -> budget
   - ยอดใช้ไป -> spent
   - ยอดคงเหลือ -> remaining
   - รวมใช้ไป -> spent
   - คงเหลือ -> remaining

4. ข้อความอื่นๆ -> แปลเป็นภาษาอังกฤษแบบสั้นกระชับ

5. ตอบเป็น JSON object เท่านั้น: {"ค่าเดิม": "ค่าใหม่", ...}
6. ถ้าเป็นตัวเลขหรือ format ถูกต้องอยู่แล้ว ไม่ต้องใส่ใน mapping"""
            },
            {
                "role": "user",
                "content": f"แปลงข้อมูลเหล่านี้:\n{json.dumps(values, ensure_ascii=False)}"
            }
        ]
    )

    result = response.choices[0].message.content.strip()
    if result.startswith("```"):
        lines = result.split("\n")
        lines = [line for line in lines if not line.startswith("```")]
        result = "\n".join(lines).strip()

    try:
        return json.loads(result)
    except Exception:
        # ถ้า JSON ไม่สมบูรณ์ ให้ return empty dict แทน crash
        print(f"   ⚠️ JSON parse failed, skipping batch")
        return {}


def analyze_dataframe_structure(df: pd.DataFrame) -> dict:
    """ให้ GPT วิเคราะห์โครงสร้าง DataFrame และบอกว่า column ไหนคืออะไร"""

    sample_data = {}
    seen_cols = set()
    for col in df.columns[:10]:
        if col in seen_cols:
            continue
        seen_cols.add(col)
        try:
            series = df[col]
            if isinstance(series, pd.DataFrame):
                series = series.iloc[:, 0]
            unique_vals = series.dropna().unique()[:5].tolist()
            sample_data[col] = [str(v) for v in unique_vals]
        except Exception:
            sample_data[col] = []

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": """วิเคราะห์โครงสร้าง DataFrame และระบุ:
1. month_column: ชื่อ column ที่เก็บเดือน (ถ้ามี)
2. type_column: ชื่อ column ที่เก็บประเภท (budget/spent/remaining) (ถ้ามี)
3. type_values: ค่าที่หมายถึง budget, spent, remaining

ตอบเป็น JSON:
{
    "month_column": "ชื่อ column หรือ null",
    "type_column": "ชื่อ column หรือ null",
    "type_values": {
        "budget": "ค่าที่หมายถึง budget หรือ null",
        "spent": "ค่าที่หมายถึง spent หรือ null",
        "remaining": "ค่าที่หมายถึง remaining หรือ null"
    }
}"""
            },
            {
                "role": "user",
                "content": f"วิเคราะห์ DataFrame นี้:\nColumns: {df.columns.tolist()}\nSample data: {json.dumps(sample_data, ensure_ascii=False, default=str)}"
            }
        ]
    )

    result = response.choices[0].message.content.strip()
    if result.startswith("```"):
        lines = result.split("\n")
        lines = [line for line in lines if not line.startswith("```")]
        result = "\n".join(lines).strip()

    return json.loads(result)


def convert_excel(excel_path: str, sheet_name: str) -> pd.DataFrame:
    """แปลง Excel ที่มี merged cells เป็น DataFrame"""

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    # เก็บค่าจาก merged cells
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_left_value

    # อ่านข้อมูลทั้งหมด
    data = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        row_data = []
        for col_idx, cell in enumerate(row, start=1):
            if (row_idx, col_idx) in merged_values:
                row_data.append(merged_values[(row_idx, col_idx)])
            else:
                row_data.append(cell.value)
        data.append(row_data)

    # สร้าง header จาก 4 rows แรก
    header_rows = data[:4]

    combined_headers = []
    for col_idx in range(len(header_rows[0])):
        parts = []
        for row in header_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip().replace('\n', ' ')
                if val and val not in parts:
                    parts.append(val)
        col_name = "_".join(parts) if parts else f"col_{col_idx}"
        combined_headers.append(col_name)

    # ทำให้ column names unique
    seen = {}
    unique_headers = []
    for h in combined_headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)

    df = pd.DataFrame(data[4:], columns=unique_headers)

    # reset index ให้สะอาดก่อน
    df = df.reset_index(drop=True)

    # dropna และ filter col_ ด้วย position แทน label เพื่อหลีก duplicate index
    keep_mask = [
        not str(c).startswith('col_') and df[c].notna().any()
        if df.columns.tolist().count(c) == 1
        else True
        for c in df.columns
    ]
    df = df.iloc[:, [i for i, k in enumerate(keep_mask) if k]]

    # deduplicate column names
    seen_d = {}
    new_cols = []
    for col in df.columns:
        if col in seen_d:
            seen_d[col] += 1
            new_cols.append(f"{col}_{seen_d[col]}")
        else:
            seen_d[col] = 0
            new_cols.append(col)
    df.columns = new_cols
    df = df.copy()  # force pandas ให้ใช้ index ใหม่สะอาด

    # ให้ GPT วิเคราะห์โครงสร้างก่อน (ใช้ชื่อ column ต้นฉบับ)
    print("🤖 GPT กำลังวิเคราะห์โครงสร้าง DataFrame...")
    structure = analyze_dataframe_structure(df)
    print(f"   📌 Month column: {structure.get('month_column')}")
    print(f"   📌 Type column: {structure.get('type_column')}")
    print(f"   📌 Type values: {structure.get('type_values')}")

    # ให้ GPT สร้าง column mapping
    print("\n🤖 GPT กำลังแปลงชื่อ columns...")
    column_mapping = get_column_mapping_from_gpt(df.columns.tolist())
    print(f"   ✅ สร้าง mapping สำหรับ {len(column_mapping)} columns")

    # ตรวจสอบ: month_column และ type_column ต้องได้ date/details
    orig_month = structure.get('month_column')
    orig_type = structure.get('type_column')
    if orig_month and orig_month in column_mapping and column_mapping[orig_month] != "date":
        print(f"   ⚠️ force: month column '{orig_month}' → 'date'")
        column_mapping[orig_month] = "date"
    if orig_type and orig_type in column_mapping and column_mapping[orig_type] != "details":
        print(f"   ⚠️ force: type column '{orig_type}' → 'details'")
        column_mapping[orig_type] = "details"

    df = df.rename(columns=column_mapping)

    # Update structure ด้วยชื่อ column ใหม่
    if orig_month and orig_month in column_mapping:
        structure['month_column'] = column_mapping[orig_month]
    if orig_type and orig_type in column_mapping:
        structure['type_column'] = column_mapping[orig_type]

    # force unique หลัง rename — GPT อาจ map หลาย col เป็นชื่อเดียวกัน
    seen_r = {}
    final_cols = []
    for col in df.columns:
        if col in seen_r:
            seen_r[col] += 1
            final_cols.append(f"{col}_{seen_r[col]}")
        else:
            seen_r[col] = 0
            final_cols.append(col)
    df.columns = final_cols
    df = df.copy()

    # แปลงข้อมูล Thai values — ใช้ iloc ตลอดเพื่อหลีก label issues
    thai_values = set()
    for i in range(len(df.columns)):
        try:
            series = df.iloc[:, i]
            for val in series.dropna().unique():
                if isinstance(val, str) and any('\u0e00' <= c <= '\u0e7f' for c in val):
                    thai_values.add(val)
        except Exception:
            pass

    if thai_values:
        print(f"\n🤖 GPT กำลังแปลงข้อมูล {len(thai_values)} รายการ...")
        data_mapping = get_data_mapping_from_gpt(list(thai_values))
        print("   ✅ แปลงข้อมูลเรียบร้อย")

        for i in range(len(df.columns)):
            df.iloc[:, i] = df.iloc[:, i].map(
                lambda x: data_mapping.get(x, x) if isinstance(x, str) else x
            )

        if structure.get('type_values'):
            for key in structure['type_values']:
                old_val = structure['type_values'][key]
                if old_val and old_val in data_mapping:
                    structure['type_values'][key] = data_mapping[old_val]

    # Filter rows ตาม structure — ใช้ iloc ตลอด
    type_col = structure.get('type_column')
    type_vals = structure.get('type_values', {})

    if type_col and type_col in df.columns.tolist():
        valid_types = [v for v in type_vals.values() if v]
        if valid_types:
            print(f"\n🔍 Filtering by {type_col} in {valid_types}")
            type_col_idx = df.columns.tolist().index(type_col)
            mask = df.iloc[:, type_col_idx].isin(valid_types)
            df = df.iloc[mask.values].copy()
            df = df.reset_index(drop=True)

    # Fill month ลงมา
    month_col = structure.get('month_column')
    if month_col and month_col in df.columns.tolist():
        month_col_idx = df.columns.tolist().index(month_col)
        filled = df.iloc[:, month_col_idx].ffill()
        df.iloc[:, month_col_idx] = filled

    df = df.reset_index(drop=True)

    # Normalize date column → YYYY-MM format
    col_list = df.columns.tolist()
    if "date" in col_list:
        date_idx = col_list.index("date")
        import re
        def _normalize_date(v):
            if not isinstance(v, str):
                return v
            # 2023-10-19 00:00:00 → 2023-10
            m = re.match(r"(\d{4})-(\d{2})", v)
            if m:
                return f"{m.group(1)}-{m.group(2)}"
            return v
        df.iloc[:, date_idx] = df.iloc[:, date_idx].map(_normalize_date)
        col_list = df.columns.tolist()

    # Post-processing — ใช้ iloc สำหรับ date/details ด้วย
    if "date" in col_list and "details" in col_list:
        date_idx = col_list.index("date")
        details_idx = col_list.index("details")
        mask_allyr = df.iloc[:, date_idx].isna() & (df.iloc[:, details_idx] == "budget")
        df.iloc[mask_allyr.values, date_idx] = "all-year-budget"

    if "date" in df.columns.tolist():
        date_idx = df.columns.tolist().index("date")
        summary_values = {"spent", "remaining", "total spent"}
        keep = ~df.iloc[:, date_idx].isin(summary_values)
        df = df.iloc[keep.values].copy()
        df = df.reset_index(drop=True)

    # ตรวจสอบว่ายังมีภาษาไทยเหลืออยู่ไหม
    remaining_thai = set()
    for i, col in enumerate(df.columns):
        if any('\u0e00' <= c <= '\u0e7f' for c in str(col)):
            remaining_thai.add(f"column: {col}")
        try:
            for val in df.iloc[:, i].dropna().unique():
                if isinstance(val, str) and any('\u0e00' <= c <= '\u0e7f' for c in val):
                    remaining_thai.add(f"data: {val}")
        except Exception:
            pass

    if remaining_thai:
        print(f"\n⚠️ ยังมีภาษาไทยเหลือ: {remaining_thai}")

    return df


if __name__ == "__main__":
    excel_path = "data/คุมฎีกาปี2566-รายได้.xlsx"
    sheet_name = "สรุปรายเดือน 66"
    csv_path = "data/finance_2026_test.csv"

    print(f"📊 Converting: {excel_path}\n")
    df = convert_excel(excel_path, sheet_name)
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"\n✅ Saved to {csv_path}")
    print(f"   Rows: {len(df)}, Columns: {len(df.columns)}")